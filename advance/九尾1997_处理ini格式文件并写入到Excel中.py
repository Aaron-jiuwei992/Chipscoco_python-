# -*- coding: utf-8 -*-
# author: jiuwei1997
# description:处理ini格式文件并写入到Excel中
# date:2021-5-5

"""
读取ini格式的文件，并创建一个excel文件，且为每个节点创建一个sheet，然后将节点下的键值写入到excel中，按照如下格式。
- 首行，字体白色 & 单元格背景色蓝色。
- 内容均居中。
- 边框。
"""
import configparser, os
from openpyxl import workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill


# 文件路径处理
base_dir = os.path.dirname(os.path.abspath(__file__))
target_excel_file_path = os.path.join(base_dir, "处理ini格式文件并写入到Excel中.xlsx")

# 准备Excel表格
wb = workbook.Workbook()
del wb["Sheet"]

# 读入ini文件，并进行解析
config = configparser.ConfigParser()
config.read("info.ini", encoding="utf-8")

# 获取所有的节点
sections = config.sections()
for section in sections:
    sheet = wb.create_sheet(section)
    # 循环节点，以获取每个节点下的键值对
    items = config.items(section)
    length_items = len(items)
    # 循环每个键值对，取出其键和值，准备写入Excel中
    for row_index, item in enumerate(items, 2):
        # 设计边框
        cell_list = sheet["A1":"B"+str(length_items+1)]
        for row in cell_list:
            for cell in row:
                cell.border = Border(
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                )

                # 设计字体对齐方式
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # 首行的内容设置
        A1, B1 = sheet["A1"], sheet["B1"]
        A1.value, B1.value = "键", "值"
        # 其余行的内容设置
        text1, text2 = item[0], item[1]
        cell1, cell2 = sheet.cell(row_index, 1), sheet.cell(row_index, 2)
        cell1.value, cell2.value = text1, text2
        # 设计首行的字体和背景色
        first_row_cell = sheet["A1": "B1"]
        for row in first_row_cell:
            for cell in row:
                cell.font = Font(name="微软雅黑", color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="6495ED")

wb.save(target_excel_file_path)




