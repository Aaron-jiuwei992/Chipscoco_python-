# -*- coding: utf-8 -*-
# author: jiuwei1997
# description:实现去网上获取指定地区的天气信息，并写入到Excel中。
# date:2021-5-4
import requests, os
from xml.etree import ElementTree as ET
from openpyxl import workbook

# 删除默认创建的sheet
wb = workbook.Workbook()
del wb["Sheet"]

# 处理文件路径
base_dir = os.path.dirname(os.path.abspath(__file__))
target_excel_file_path = os.path.join(base_dir, "九尾1997_获取指定地区天气信息.xlsx")

while True:
    city = input("请输入城市（Q/q退出）：")
    if city.upper() == "Q":
        break
    url = "http://ws.webxml.com.cn//WebServices/WeatherWebService.asmx/getWeatherbyCityName?theCityName={}".format(city)
    res = requests.get(url=url)
    res.encoding = "utf-8"
    print(res.text)
    # 1.提取XML格式中的数据
    root = ET.XML(res.text)

    # 为每个城市创建一个sheet
    sheet = wb.create_sheet(city)
    for row_index, child in enumerate(root, 1):
        text = child.text
        cell = sheet.cell(row_index, 1)
        cell.value = text

    wb.save(target_excel_file_path)




